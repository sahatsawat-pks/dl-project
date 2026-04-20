"""
experiments_log.py — Manages experiments.xlsx log file.

Creates and updates an Excel file matching the course example format:
  No | Model Name | [Hyperparameters: batch_size, epochs, lr] |
  [Performance: Train Acc, Test Acc] | Overfit/Underfit/Just Right
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

XLSX_PATH = "experiments.xlsx"

# ── Styling constants (matching the green-on-dark course table) ───────────────
HEADER_BG_TOP    = "1F3864"   # dark navy — top merged headers
HEADER_BG_HYPER  = "843C0C"   # dark red  — Hyperparameters
HEADER_BG_PERF   = "375623"   # dark green — Performance
HEADER_BG_OVR    = "375623"   # dark green — Overfit column
HEADER_FG        = "FFFFFF"   # white text

COL_BG_ALT       = "F2F2F2"   # light grey alternating rows
COL_BG_O         = "FFC7CE"   # light red  — Overfit
COL_BG_B         = "C6EFCE"   # light green — Just Right
COL_BG_U         = "FFEB9C"   # light yellow — Underfit

# Column layout:  A        B           C           D        E      F             G            H
COL_NAMES = ["No", "Model Name", "batch_size", "epochs", "learning_rate",
             "Train Accuracy", "Test Accuracy",
             "Overfit (O) / Underfit (U) / Just Right (B)"]


def _thin_border():
    thin = Side(border_style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_cell(ws, row, col, value, bg, fg="FFFFFF", bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(color=fg, bold=bold, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    return cell


def _create_header(ws):
    """Write the two-row merged header matching the course example."""
    # ── Row 1: group labels ───────────────────────────────────────────────────
    # Col A: No (merged rows 1-2)
    _header_cell(ws, 1, 1, "No", HEADER_BG_TOP)
    ws.merge_cells("A1:A2")

    # Col B: Model Name (merged rows 1-2)
    _header_cell(ws, 1, 2, "Model Name", HEADER_BG_TOP)
    ws.merge_cells("B1:B2")

    # Col C-E: Hyperparameters (merged)
    _header_cell(ws, 1, 3, "Hyperparameters", HEADER_BG_HYPER)
    ws.merge_cells("C1:E1")

    # Col F-G: Performance (merged)
    _header_cell(ws, 1, 6, "Performance", HEADER_BG_PERF)
    ws.merge_cells("F1:G1")

    # Col H: Overfit label (merged rows 1-2)
    _header_cell(ws, 1, 8, "Overfit (O) / Underfit (U) / Just Right (B)",
                 HEADER_BG_OVR)
    ws.merge_cells("H1:H2")

    # ── Row 2: sub-column labels ──────────────────────────────────────────────
    for col, label in [(3, "batch_size"), (4, "epochs"), (5, "learning_rate")]:
        _header_cell(ws, 2, col, label, HEADER_BG_HYPER)

    for col, label in [(6, "Train Accuracy"), (7, "Test Accuracy")]:
        _header_cell(ws, 2, col, label, HEADER_BG_PERF)

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [5, 22, 12, 8, 14, 16, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18


def _classify(train_acc: float, test_acc: float) -> str:
    """Determine Overfit / Underfit / Just Right."""
    gap = train_acc - test_acc
    if gap > 15.0:
        return "O"   # Overfit
    elif test_acc < 55.0 and train_acc < 60.0:
        return "U"   # Underfit (both low)
    else:
        return "B"   # Just Right / Balanced


def _row_fill(label: str) -> PatternFill:
    color = {"O": COL_BG_O, "U": COL_BG_U, "B": COL_BG_B}.get(label, "FFFFFF")
    return PatternFill("solid", fgColor=color)


def log_experiment(
    model_name: str,
    batch_size:  int,
    epochs_run:  int,
    learning_rate: float,
    train_accuracy: float,   # 0–100 %
    test_accuracy:  float,   # 0–100 %
    xlsx_path: str = XLSX_PATH,
):
    """Append one row to experiments.xlsx (creates file + header if needed)."""

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiments"
        _create_header(ws)

    # Next row number (data starts at row 3)
    next_row = ws.max_row + 1
    if next_row < 3:
        next_row = 3

    entry_no = next_row - 2   # sequential experiment number (1-based from 0)

    label = _classify(train_accuracy, test_accuracy)
    fill  = _row_fill(label)

    values = [
        entry_no - 1,        # No (0-indexed like example)
        model_name,
        batch_size,
        epochs_run,
        f"{learning_rate:.2e}",
        round(train_accuracy, 1),
        round(test_accuracy, 1),
        label,
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font      = Font(size=10)
        # Colour last column by O/U/B
        if col == 8:
            cell.fill = fill
            cell.font = Font(bold=True, size=10)
        elif next_row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=COL_BG_ALT)

    wb.save(xlsx_path)
    print(f"[experiments] Logged row {entry_no-1}: {model_name} | "
          f"Train={train_accuracy:.1f}% Test={test_accuracy:.1f}% → {label}")
    return xlsx_path
